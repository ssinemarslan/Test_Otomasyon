from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec 
import pytest
import openpyxl
from constantss import globalConstants1 as c 


class Test_DemoClass:
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window()

    def teardown_method(self):
        self.driver.quit()

    def getDATA():
        excel=openpyxl.load_workbook(c.invalid_login_xlsx)
        sheet=excel["Sayfa1"]
        rows=sheet.max_row
        data=[]

        for i in range(2,rows+1):
            username=sheet.cell(i,1).value
            password=sheet.cell(i,2).value
            data.append((username,password))

        return data

    @pytest.mark.parametrize("username,password",getDATA())
    def test_unsuccesfull_login(self,username,password):
        usernameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton=self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage=self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH)
        assert errorMessage.text==c.USERNAME_PASSWORD_DONT_MATCH
    